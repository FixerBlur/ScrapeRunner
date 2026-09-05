import csv
import json
from pathlib import Path

from openpyxl import load_workbook

from scraperunner.exporter import ExportWriter, export_items_csv, export_items_xlsx, read_items_json, safe_cell
from scraperunner.models import Item, PageResult

ITEM = Item(title="Pan", link="https://s.com/p/1", image="https://s.com/i.jpg", price="449 ₴", old_price="1 038 ₴", text="Pan 449 ₴")
PAGE = PageResult(url="https://s.com/", status=200, title="Shop", links=["https://s.com/a"], images=["https://s.com/i.jpg"], items=[ITEM])


def test_writer_streams_pages_and_writes_items(tmp_path: Path):
    with ExportWriter(tmp_path) as writer:
        writer.add(PAGE)
        writer.add(PageResult(url="https://s.com/2", status=404, error="not found"))

    pages = json.loads((tmp_path / "pages.json").read_text(encoding="utf-8"))
    assert [page["url"] for page in pages] == ["https://s.com/", "https://s.com/2"]
    assert writer.stats.as_dict() == {"pages_done": 2, "failed": 1, "items": 1, "links": 1, "images": 1}
    assert writer.item_images == ["https://s.com/i.jpg"]

    links = list(csv.reader((tmp_path / "links.csv").open(encoding="utf-8")))
    assert links[0] == ["page_url", "page_title", "type", "item_url"]
    assert len(links) == 3  # header + one link + one image

    rows = json.loads((tmp_path / "items.json").read_text(encoding="utf-8"))
    assert rows[0]["title"] == "Pan" and rows[0]["page"] == "https://s.com/"
    assert read_items_json(tmp_path / "items.json") == [ITEM]
    assert (tmp_path / "items.xlsx").exists() and (tmp_path / "items.csv").exists()


def test_items_xlsx_has_header_rows_and_links(tmp_path: Path):
    path = tmp_path / "items.xlsx"
    export_items_xlsx([{**ITEM.__dict__, "page": "https://s.com/"}], path)

    sheet = load_workbook(path).active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == ("title", "price", "old_price", "link", "image", "page", "group", "text")
    assert rows[1] == ("Pan", "449 ₴", "1 038 ₴", "https://s.com/p/1", "https://s.com/i.jpg", "https://s.com/", 1, "Pan 449 ₴")
    assert sheet["D2"].hyperlink.target == "https://s.com/p/1"
    assert sheet.freeze_panes == "A2"


def test_formula_like_text_is_neutralised(tmp_path: Path):
    evil = {**ITEM.__dict__, "title": "=HYPERLINK(\"http://evil\")", "text": "+1", "page": "https://s.com/"}
    export_items_csv([evil], tmp_path / "items.csv")
    export_items_xlsx([evil], tmp_path / "items.xlsx")

    csv_row = list(csv.reader((tmp_path / "items.csv").open(encoding="utf-8")))[1]
    assert csv_row[0].startswith("'=") and csv_row[-1] == "'+1"
    assert load_workbook(tmp_path / "items.xlsx").active["A2"].value.startswith("'=")
    assert safe_cell("normal") == "normal" and safe_cell(5) == 5 and safe_cell("-50% sale") == "'-50% sale"
