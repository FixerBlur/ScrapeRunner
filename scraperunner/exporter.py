from __future__ import annotations

import csv
import json
from collections.abc import Sequence
from dataclasses import asdict, dataclass, field, fields
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from scraperunner.models import Item, PageResult

if TYPE_CHECKING:
    from scraperunner.compare import Changes

ITEM_COLUMNS = ("title", "price", "old_price", "link", "image", "page", "group", "text")
LINK_COLUMNS = ("page_url", "page_title", "type", "item_url")
_XLSX_WIDTHS = {"title": 50, "price": 12, "old_price": 12, "link": 45, "image": 45, "page": 35, "group": 8, "text": 80}
_URL_COLUMNS = {"link", "image", "page"}
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")
_RECENT = 8


def safe_cell(value):
    """Neutralise spreadsheet formula injection: scraped text must never start with = + - @."""
    return f"'{value}" if isinstance(value, str) and value.startswith(_FORMULA_PREFIXES) else value


@dataclass
class CrawlStats:
    """Running totals of a crawl, small enough to keep for any number of pages."""

    pages: int = 0
    failed: int = 0
    items: int = 0
    links: int = 0
    image_urls: dict[str, None] = field(default_factory=dict)  # insertion-ordered set
    recent: list[str] = field(default_factory=list)

    def add(self, page: PageResult) -> None:
        self.pages += 1
        self.failed += bool(page.error)
        self.items += len(page.items)
        self.links += len(page.links)
        self.image_urls.update(dict.fromkeys(page.images))
        self.recent = [*self.recent, page.url][-_RECENT:]

    @property
    def images(self) -> list[str]:
        return list(self.image_urls)

    def as_dict(self) -> dict:
        return {"pages_done": self.pages, "failed": self.failed, "items": self.items, "links": self.links, "images": len(self.image_urls)}


class ExportWriter:
    """Streams pages to disk as they arrive, so a long crawl never holds every page in memory.

    ``pages.json`` and ``links.csv`` grow page by page; items are kept as flat
    rows (a few hundred bytes each) and written as xlsx/csv/json on ``close``.
    """

    def __init__(self, output_dir: Path) -> None:
        output_dir.mkdir(parents=True, exist_ok=True)
        self.output_dir = output_dir
        self.stats = CrawlStats()
        self.item_rows: list[dict] = []
        self._item_images: dict[str, None] = {}
        self._pages = (output_dir / "pages.json").open("w", encoding="utf-8")
        self._pages.write("[\n")
        self._links_file = (output_dir / "links.csv").open("w", encoding="utf-8", newline="")
        self._links = csv.writer(self._links_file)
        self._links.writerow(LINK_COLUMNS)
        self._closed = False

    def add(self, page: PageResult) -> None:
        self._pages.write(("" if self.stats.pages == 0 else ",\n") + json.dumps(page.to_dict(), ensure_ascii=False))
        for kind, urls in (("link", page.links), ("image", page.images)):
            self._links.writerows((safe_cell(page.url), safe_cell(page.title), kind, safe_cell(url)) for url in urls)
        for item in page.items:
            self.item_rows.append({**asdict(item), "page": page.url})
            if item.image:
                self._item_images[item.image] = None
        self.stats.add(page)

    @property
    def item_images(self) -> list[str]:
        return list(self._item_images)

    def close(self) -> None:
        if self._closed:
            return
        self._closed = True
        self._pages.write("\n]\n")
        self._pages.close()
        self._links_file.close()
        export_items_json(self.item_rows, self.output_dir / "items.json")
        export_items_csv(self.item_rows, self.output_dir / "items.csv")
        export_items_xlsx(self.item_rows, self.output_dir / "items.xlsx")

    def __enter__(self) -> "ExportWriter":
        return self

    def __exit__(self, *exc) -> None:
        self.close()


def export_items_json(rows: Sequence[dict], path: Path) -> None:
    _write_json(list(rows), path)


def export_items_csv(rows: Sequence[dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(ITEM_COLUMNS)
        writer.writerows(tuple(safe_cell(row[column]) for column in ITEM_COLUMNS) for row in rows)


def export_items_xlsx(rows: Sequence[dict], path: Path) -> None:
    """Excel sheet: bold frozen header, filters, sensible widths, clickable URLs."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Items"
    sheet.append(list(ITEM_COLUMNS))
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        sheet.append([safe_cell(row[column]) for column in ITEM_COLUMNS])
        for cell, column in zip(sheet[sheet.max_row], ITEM_COLUMNS):
            if column in _URL_COLUMNS and cell.value:
                cell.hyperlink = cell.value
                cell.font = Font(color="0563C1", underline="single")

    for index, column in enumerate(ITEM_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = _XLSX_WIDTHS[column]
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def export_changes_json(changes: "Changes", path: Path) -> None:
    _write_json(changes.to_dict(), path)


def read_items_json(path: Path) -> list[Item]:
    """Items back from an ``items.json``; the extra ``page`` column is dropped."""
    with path.open(encoding="utf-8") as fh:
        rows = json.load(fh)
    return rows_to_items(rows)


def rows_to_items(rows: Sequence[dict]) -> list[Item]:
    names = {f.name for f in fields(Item)}
    return [Item(**{key: value for key, value in row.items() if key in names}) for row in rows]


def _write_json(data, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=False, indent=2)
